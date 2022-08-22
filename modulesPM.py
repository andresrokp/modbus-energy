
from genericpath import exists
from pymodbus.client.sync import ModbusTcpClient
import openpyxl
import datetime

def pickAddress(ipAdrs):
    ip = ipAdrs[0]
    name =  ipAdrs[1]
    return ip, name

def queryPmData(ipAdrs):
    client = ModbusTcpClient(ipAdrs, timeout=2)
    # registers are addressed starting from zero, 
    # so we need to subtract 1 from the address
    # energy data is stored in two consecutive registers
    unit=0x02 if ["192.168.20.45", "192.168.25.10"].count(ipAdrs) else 0x00
    result = client.read_holding_registers(2699, 2, unit=unit)
    client.close()
    return conv754toDEC(*result.registers)

def conv754toDEC(msw, lsw):
    denom = 0x800000
    mantissa = 0
    expon = 0
    for i in range(16,0,-1):
        if lsw & 1:
            mantissa += (1 / denom)
        denom = denom >> 1
        lsw = lsw >> 1
    for i in range(7,0,-1):
        if msw & 1:
            mantissa += (1 / denom)
        denom = denom >> 1
        msw = msw >> 1
    expon = (0xff & msw) - 127
    if msw & 0x100:
        sign = -1
    else:
        sign = 1
    return (sign * (1 + mantissa) * 2**expon);

def consolePrint(result, ip, name):
    print(ip, "\n", name)
    if type(result) == float:
        print(" Pot Activa Acum:  ", result,"kWh")
        print("-")
    else:
        print(" :(  Error, :: ", result)
        print("-")

def xlxsPrint(dataLogger):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([" "])
    ws.merge_cells('A2:C2')
    ws['A2'] = "REGISTRO DE CONSUMO ACUMULADO MEDIDORES DE ENERGÍA"
    moment = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws.append(['Momento captura: '+moment])
    ws.append([" "])
    ws.append([" "])
    
    ws.append(["Nombre", "IP", "Pot Activa Acum"])
    ws.append([" "])
    for log in dataLogger:
        print(log)
        ws.append(log)
    
    fileFriendlyMoment = moment.replace(":", "-").replace("/", "-")
    wb.save(f'PMs {fileFriendlyMoment}.xlsx')
    wb.close()